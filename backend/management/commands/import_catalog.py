"""
Import a catalog Excel file into the database.

Usage:
    python manage.py import_catalog path/to/COSWELL_HF_Gros.xlsx --org 1 --supplier 5

The Excel file must follow the standard 10-column format:
    Row 1-2: Header (Fornitore in C2, Cliente in E2)
    Row 3: Column headers (Color, Codice, EAN, CodIntFor, Descrizione, Peso, QxC, Promo, Ordine, Omaggi)
    Row 4+: Category separators (color in A, name in E, no EAN) and product rows

Can also bulk-import all .xlsx files in a directory:
    python manage.py import_catalog path/to/folder/ --org 1 --auto-match
"""

import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Import catalog Excel file(s) into the PWA catalog tables"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            help="Path to an .xlsx file or a directory containing .xlsx files",
        )
        parser.add_argument(
            "--org",
            type=int,
            required=True,
            help="Organization ID",
        )
        parser.add_argument(
            "--supplier",
            type=int,
            default=None,
            help="Supplier (Fornitore) ID. Required for single file import.",
        )
        parser.add_argument(
            "--auto-match",
            action="store_true",
            help="Auto-match supplier from filename (e.g. COSWELL_HF_Gros.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and show what would be imported, without writing to DB",
        )

    def handle(self, *args, **options):
        path = options["path"]
        org_id = options["org"]

        # Import models here to avoid issues with --skip-checks
        from api.models import Fornitore  # Adjust import to your app
        from api.models_pwa import Catalog, CatalogCategory, CatalogProduct

        try:
            from api.models import Organization

            org = Organization.objects.get(pk=org_id)
        except Exception as e:
            raise CommandError(f"Organization {org_id} not found: {e}")

        # Collect files to process
        if os.path.isdir(path):
            files = sorted(
                [
                    os.path.join(path, f)
                    for f in os.listdir(path)
                    if f.endswith(".xlsx") and not f.startswith("~")
                ]
            )
            self.stdout.write(f"Found {len(files)} Excel files in {path}")
        elif os.path.isfile(path) and path.endswith(".xlsx"):
            files = [path]
        else:
            raise CommandError(f"Invalid path: {path}")

        total_products = 0

        for filepath in files:
            filename = os.path.basename(filepath)
            self.stdout.write(f"\n{'='*60}")
            self.stdout.write(f"Processing: {filename}")

            # Parse the Excel file
            parsed = self.parse_excel(filepath)
            if not parsed:
                self.stderr.write(f"  SKIP: Could not parse {filename}")
                continue

            # Resolve supplier
            supplier = None
            if options["supplier"]:
                try:
                    supplier = Fornitore.objects.get(pk=options["supplier"])
                except Fornitore.DoesNotExist:
                    raise CommandError(
                        f"Supplier ID {options['supplier']} not found"
                    )
            elif options["auto_match"]:
                supplier = self.auto_match_supplier(filename, Fornitore)

            if not supplier:
                self.stderr.write(
                    f"  SKIP: No supplier for {filename}. "
                    f"Use --supplier or --auto-match."
                )
                continue

            # Extract client name from filename or header
            client_name = parsed.get("client_name") or self.extract_client_from_filename(filename)
            if not client_name:
                self.stderr.write(f"  SKIP: Cannot determine client for {filename}")
                continue

            self.stdout.write(
                f"  Supplier: {supplier.Title}"
                f"\n  Client: {client_name}"
                f"\n  Categories: {len(parsed['categories'])}"
                f"\n  Products: {sum(len(c['products']) for c in parsed['categories'])}"
            )

            if options["dry_run"]:
                for cat in parsed["categories"]:
                    self.stdout.write(
                        f"    [{cat['color']}] {cat['name']}: "
                        f"{len(cat['products'])} products"
                    )
                continue

            # Write to database
            catalog, created = Catalog.objects.update_or_create(
                organization=org,
                supplier=supplier,
                client_name=client_name,
                defaults={
                    "client_code": client_name.upper(),
                    "source_file": filename,
                    "is_active": True,
                },
            )

            if not created:
                # Clear existing categories and products for re-import
                catalog.categories.all().delete()
                self.stdout.write(f"  Updated existing catalog (cleared old data)")
            else:
                self.stdout.write(f"  Created new catalog")

            # Create categories and products
            for cat_order, cat_data in enumerate(parsed["categories"]):
                category = CatalogCategory.objects.create(
                    catalog=catalog,
                    name=cat_data["name"],
                    color=cat_data["color"],
                    sort_order=cat_order,
                )

                products_to_create = []
                for prod_order, prod_data in enumerate(cat_data["products"]):
                    products_to_create.append(
                        CatalogProduct(
                            category=category,
                            catalog=catalog,
                            code=prod_data.get("code", ""),
                            ean=prod_data.get("ean", ""),
                            internal_code=prod_data.get("internal_code", ""),
                            description=prod_data["description"],
                            weight=prod_data.get("weight", ""),
                            qxc=prod_data.get("qxc", 1),
                            default_promo=prod_data.get("promo", ""),
                            sort_order=prod_order,
                        )
                    )

                CatalogProduct.objects.bulk_create(products_to_create)
                total_products += len(products_to_create)

        self.stdout.write(
            self.style.SUCCESS(
                f"\nDone! Imported {total_products} products from {len(files)} file(s)."
            )
        )

    def parse_excel(self, filepath):
        """
        Parse a catalog Excel file into a structured dict.
        Returns: {
            "supplier_name": "COSWELL HF",
            "client_name": "Gros",
            "categories": [
                {
                    "name": "Tisane a caldo",
                    "color": "#7DEC18",
                    "products": [
                        { "code": "GA1992900", "ean": "...", "description": "...", ... },
                    ]
                }
            ]
        }
        """
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stderr.write(f"  Error opening file: {e}")
            return None

        # Read header
        supplier_name = str(ws.cell(row=2, column=3).value or "").strip()
        client_header = str(ws.cell(row=2, column=5).value or "").strip()

        # Parse rows from row 4 onward
        categories = []
        current_category = None

        for row_idx in range(4, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value  # Color code
            b_val = ws.cell(row=row_idx, column=2).value  # Codice
            c_val = ws.cell(row=row_idx, column=3).value  # EAN
            d_val = ws.cell(row=row_idx, column=4).value  # CodIntFor
            e_val = ws.cell(row=row_idx, column=5).value  # Descrizione
            f_val = ws.cell(row=row_idx, column=6).value  # Peso
            g_val = ws.cell(row=row_idx, column=7).value  # QxC
            h_val = ws.cell(row=row_idx, column=8).value  # Promo

            if not e_val:
                continue

            # Detect category row: has color in A + text in E, but no EAN or CodIntFor
            is_category = bool(a_val and e_val and not c_val and not d_val)

            if is_category:
                # Try to extract color from cell fill
                cell_a = ws.cell(row=row_idx, column=1)
                color = self.extract_color(cell_a, a_val)

                current_category = {
                    "name": str(e_val).strip(),
                    "color": color,
                    "products": [],
                }
                categories.append(current_category)
            else:
                # Product row
                if current_category is None:
                    # Products before any category — create a default one
                    current_category = {
                        "name": "Generale",
                        "color": "#10B981",
                        "products": [],
                    }
                    categories.append(current_category)

                # Parse QxC — handle string values like "10" or empty
                qxc = 1
                if g_val:
                    try:
                        qxc = int(float(str(g_val).strip()))
                    except (ValueError, TypeError):
                        qxc = 1

                current_category["products"].append(
                    {
                        "code": str(b_val).strip() if b_val else "",
                        "ean": str(c_val).strip() if c_val else "",
                        "internal_code": str(d_val).strip() if d_val else "",
                        "description": str(e_val).strip(),
                        "weight": str(f_val).strip() if f_val else "",
                        "qxc": qxc,
                        "promo": str(h_val).strip() if h_val else "",
                    }
                )

        wb.close()

        # Try to extract client name from header
        client_name = ""
        if client_header:
            # Header format is usually "Client/via - Agent" or just "Client"
            parts = client_header.split("/")
            client_name = parts[0].strip() if parts else client_header.strip()

        return {
            "supplier_name": supplier_name,
            "client_name": client_name,
            "categories": categories,
        }

    def extract_color(self, cell, text_value):
        """Try to extract a hex color from cell fill or text value."""
        # Try cell fill color
        try:
            if cell.fill and cell.fill.start_color:
                rgb = cell.fill.start_color.rgb
                if rgb and rgb != "00000000" and len(rgb) >= 6:
                    # Handle ARGB format (8 chars) or RGB (6 chars)
                    hex_color = rgb[-6:]
                    return f"#{hex_color}"
        except Exception:
            pass

        # Check if text_value itself is a hex color
        if isinstance(text_value, str) and text_value.startswith("#"):
            return text_value

        return "#10B981"  # Default brand green

    def auto_match_supplier(self, filename, FornitoreModel):
        """
        Try to match a supplier from the filename.
        Filename pattern: BRAND_Client_hash.xlsx
        e.g. COSWELL_HF_Gros_lRN5pPv.xlsx → look for "COSWELL HF"
        """
        name = filename.replace(".xlsx", "")
        parts = name.split("_")

        # Try progressively shorter prefixes
        # COSWELL_HF_Gros_hash → try "COSWELL HF", then "COSWELL"
        for i in range(len(parts) - 1, 0, -1):
            candidate = " ".join(parts[:i])
            supplier = FornitoreModel.objects.filter(
                Title__icontains=candidate
            ).first()
            if supplier:
                return supplier

        return None

    def extract_client_from_filename(self, filename):
        """
        Extract client name from filename pattern.
        COSWELL_HF_Gros_hash.xlsx → "Gros"
        DC_Pac2000_hash.xlsx → "Pac2000"
        """
        name = filename.replace(".xlsx", "")
        parts = name.split("_")

        # Known client names
        known_clients = {"gros", "elite", "pac2000"}

        for part in parts:
            if part.lower() in known_clients:
                return part

        # If not found, try second-to-last part (before hash)
        if len(parts) >= 3:
            return parts[-2]

        return None
